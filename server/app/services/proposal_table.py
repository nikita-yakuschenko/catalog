"""Extract estimate rows as [title, price] pairs — one row, one match.

Primary path for KP intake. MarkItDown split-columns is only a fallback.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from app.services.proposal_parse import document_from_table_rows

logger = logging.getLogger(__name__)

_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_PDF_SUFFIXES = {".pdf"}

# pdfplumber: линии сетки, затем текст (сметы из Excel часто без нормальных линий)
_PDFPLUMBER_STRATEGIES = (
    {},  # default = lines
    {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_tolerance": 3,
        "intersection_tolerance": 3,
    },
)


def extract_excel_rows(path: Path) -> list[list[Any]]:
    """Read first sheet row-by-row via openpyxl (keeps title↔price cells together)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            values = list(row)
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            rows.append(values)
        return rows
    finally:
        wb.close()


def extract_pdf_rows_pymupdf(path: Path) -> list[list[Any]]:
    """PyMuPDF find_tables: сначала lines, затем text — взять лучший результат."""
    import fitz

    candidates: list[list[list[Any]]] = []
    doc = fitz.open(path)
    try:
        for strategy in ("lines", "text"):
            rows: list[list[Any]] = []
            for page in doc:
                try:
                    finder = page.find_tables(strategy=strategy)
                except TypeError:
                    # Старые сигнатуры без strategy
                    try:
                        finder = page.find_tables()
                    except Exception as exc:
                        logger.warning("pymupdf find_tables failed: %s", exc)
                        continue
                except Exception as exc:
                    logger.warning("pymupdf find_tables(%s) failed: %s", strategy, exc)
                    continue
                tables = getattr(finder, "tables", None) or []
                for table in tables:
                    try:
                        extracted = table.extract() or []
                    except Exception as exc:
                        logger.warning("pymupdf table.extract failed: %s", exc)
                        continue
                    for row in extracted:
                        if row and any(cell not in (None, "") for cell in row):
                            rows.append(list(row))
            if rows:
                candidates.append(rows)
            if strategy == "lines" and _score_rows(rows) >= 50:
                # Достаточно богатая таблица по линиям — text не нужен
                break
    finally:
        doc.close()
    return _best_rows(candidates)


def extract_pdf_rows_pdfplumber(path: Path) -> list[list[Any]]:
    """pdfplumber extract_tables с несколькими стратегиями."""
    import pdfplumber

    candidates: list[list[list[Any]]] = []
    with pdfplumber.open(path) as pdf:
        for settings in _PDFPLUMBER_STRATEGIES:
            rows: list[list[Any]] = []
            for page in pdf.pages:
                try:
                    tables = (
                        page.extract_tables(table_settings=settings)
                        if settings
                        else page.extract_tables()
                    ) or []
                except Exception as exc:
                    logger.warning("pdfplumber extract_tables failed: %s", exc)
                    continue
                for table in tables:
                    for row in table or []:
                        if row and any(cell not in (None, "") for cell in row):
                            rows.append(list(row))
            if rows:
                candidates.append(rows)
            if settings == {} and _score_rows(rows) >= 50:
                break
    return _best_rows(candidates)


def _score_document(doc: dict[str, Any]) -> int:
    """Higher = better structured estimate."""
    options = doc.get("options") or []
    priced = sum(1 for o in options if o.get("price"))
    score = priced * 10
    if doc.get("house_price"):
        score += 50
    if doc.get("project_name"):
        score += 5
    # Штраф, если ИТОГО просочился в опции (не должно после фильтров)
    for opt in options:
        title = str(opt.get("title") or "").lower()
        if title.startswith("итог") or title.startswith("всего"):
            score -= 100
    return score


def _score_rows(rows: list[list[Any]]) -> int:
    if not rows:
        return -1
    return _score_document(document_from_table_rows(rows))


def _best_rows(candidates: list[list[list[Any]]]) -> list[list[Any]]:
    best_rows: list[list[Any]] = []
    best_score = -1
    for rows in candidates:
        if not rows:
            continue
        score = _score_rows(rows)
        if score > best_score:
            best_score = score
            best_rows = rows
    return best_rows


def extract_pdf_table_rows(path: Path) -> tuple[list[list[Any]], str]:
    """PyMuPDF и pdfplumber — вернуть лучшие строки и имя движка."""
    scored: list[tuple[int, str, list[list[Any]]]] = []

    try:
        rows = extract_pdf_rows_pymupdf(path)
        if rows:
            scored.append((_score_rows(rows), "pdf_pymupdf", rows))
    except Exception as exc:
        logger.warning("pymupdf table path failed: %s", exc)

    try:
        rows = extract_pdf_rows_pdfplumber(path)
        if rows:
            scored.append((_score_rows(rows), "pdf_pdfplumber", rows))
    except Exception as exc:
        logger.warning("pdfplumber table path failed: %s", exc)

    if not scored:
        return [], ""

    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, engine, best_rows = scored[0]
    logger.info(
        "pdf table extract engine=%s score=%s rows=%s",
        engine,
        best_score,
        len(best_rows),
    )
    return best_rows, engine


def rows_look_usable(rows: list[list[Any]]) -> bool:
    if not rows:
        return False
    doc = document_from_table_rows(rows)
    return bool(doc.get("house_price") or doc.get("options"))


def extract_estimate_document(path: Path) -> tuple[dict[str, Any], str, str]:
    """Parse estimate file into KP document.

    Returns (document, intake_trace, method) where method is
    excel | pdf_pymupdf | pdf_pdfplumber | markdown_fallback.
    """
    path = path.resolve()
    suffix = path.suffix.lower()

    if suffix in _EXCEL_SUFFIXES:
        try:
            rows = extract_excel_rows(path)
            if rows_look_usable(rows):
                doc = document_from_table_rows(rows)
                return doc, _rows_trace(rows), "excel"
        except Exception as exc:
            logger.warning("excel extract failed %s: %s", path.name, exc)

    if suffix in _PDF_SUFFIXES:
        try:
            rows, engine = extract_pdf_table_rows(path)
            if rows_look_usable(rows):
                doc = document_from_table_rows(rows)
                return doc, _rows_trace(rows), engine or "pdf_table"
        except Exception as exc:
            logger.warning("pdf table extract failed %s: %s", path.name, exc)

    # Fallback: MarkItDown / plain text (may mis-pair split columns)
    from app.services.proposal_intake import file_to_markdown
    from app.services.proposal_parse import parse_markdown

    markdown = file_to_markdown(path)
    doc = parse_markdown(markdown)
    return doc, markdown, "markdown_fallback"


def _rows_trace(rows: list[list[Any]]) -> str:
    lines: list[str] = []
    for row in rows:
        cells = ["" if c is None else str(c).strip() for c in row]
        lines.append(" | ".join(cells))
    return "\n".join(lines)
